from __future__ import annotations

import csv
import io
import re
import unicodedata
import urllib.request
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from .domain import Draw, PrizeResult

OFFICIAL_HISTORY_PAGE = "https://www.fdj.fr/jeux-de-tirage/loto/historique"
LATEST_ARCHIVE_URL = (
    "https://www.sto.api.fdj.fr/anonymous/service-draw-info/v3/documentations/"
    "1a2b3c4d-9876-4562-b3fc-2c963f66afp6"
)
OFFICIAL_ARCHIVES = {
    "loto-1976-2008.zip": (
        "https://www.sto.api.fdj.fr/anonymous/service-draw-info/v3/documentations/"
        "1a2b3c4d-9876-4562-b3fc-2c963f66afl6"
    ),
    "loto-2008-2017.zip": (
        "https://www.sto.api.fdj.fr/anonymous/service-draw-info/v3/documentations/"
        "1a2b3c4d-9876-4562-b3fc-2c963f66afm6"
    ),
    "loto-2017-2019.zip": (
        "https://www.sto.api.fdj.fr/anonymous/service-draw-info/v3/documentations/"
        "1a2b3c4d-9876-4562-b3fc-2c963f66afn6"
    ),
    "loto-2019a.zip": (
        "https://www.sto.api.fdj.fr/anonymous/service-draw-info/v3/documentations/"
        "1a2b3c4d-9876-4562-b3fc-2c963f66afo6"
    ),
    "loto-latest.zip": LATEST_ARCHIVE_URL,
    "grand-loto-2017-2018.zip": (
        "https://www.sto.api.fdj.fr/anonymous/service-draw-info/v3/documentations/"
        "1a2b3c4d-9876-4562-b3fc-2c963f66aff6"
    ),
    "grand-loto-2019-2025.zip": (
        "https://www.sto.api.fdj.fr/anonymous/service-draw-info/v3/documentations/"
        "1a2b3c4d-9876-4562-b3fc-2c963f66afg6"
    ),
    "super-loto-1996-2008.zip": (
        "https://www.sto.api.fdj.fr/anonymous/service-draw-info/v3/documentations/"
        "1a2b3c4d-9876-4562-b3fc-2c963f66afh6"
    ),
    "super-loto-2009-2017.zip": (
        "https://www.sto.api.fdj.fr/anonymous/service-draw-info/v3/documentations/"
        "1a2b3c4d-9876-4562-b3fc-2c963f66afi6"
    ),
    "super-loto-2017-2018.zip": (
        "https://www.sto.api.fdj.fr/anonymous/service-draw-info/v3/documentations/"
        "1a2b3c4d-9876-4562-b3fc-2c963f66afj6"
    ),
    "super-loto-2019-2026.zip": (
        "https://www.sto.api.fdj.fr/anonymous/service-draw-info/v3/documentations/"
        "1a2b3c4d-9876-4562-b3fc-2c963f66afk6"
    ),
}


@dataclass(frozen=True, slots=True)
class LegacyDraw:
    main: tuple[int, ...]
    complementary: int
    draw_date: date | None = None
    game: str = "loto"
    prizes: tuple[PrizeResult, ...] = ()

    def __post_init__(self) -> None:
        normalized = tuple(sorted(self.main))
        if len(normalized) != 6 or len(set(normalized)) != 6:
            raise ValueError("Un ancien tirage doit contenir 6 numeros distincts")
        if normalized[0] < 1 or normalized[-1] > 49:
            raise ValueError("Les anciens numeros doivent etre compris entre 1 et 49")
        if not 1 <= self.complementary <= 49 or self.complementary in normalized:
            raise ValueError("Le numero complementaire historique est invalide")
        if not self.game:
            raise ValueError("Le type de jeu ne peut pas etre vide")
        object.__setattr__(self, "main", normalized)


def _infer_game(name: str | Path) -> str:
    normalized = str(name).lower()
    if "super" in normalized or "sloto" in normalized:
        return "super_loto"
    if "grand" in normalized or "noel" in normalized:
        return "grand_loto"
    return "loto"


def _normalize(value: str) -> str:
    plain = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "_", plain.lower()).strip("_")


def _parse_date(value: str) -> datetime | None:
    value = value.strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y%m%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    return None


def _find_column(headers: dict[str, str], aliases: Iterable[str]) -> str | None:
    for alias in aliases:
        if alias in headers:
            return headers[alias]
    return None


def _optional_int(value: object) -> int | None:
    normalized = str(value).strip().replace(" ", "")
    if not normalized or normalized.lower() == "none":
        return None
    return int(float(normalized.replace(",", ".")))


def _optional_float(value: object) -> float | None:
    normalized = str(value).strip().replace(" ", "")
    if not normalized or normalized.lower() == "none":
        return None
    return float(normalized.replace(",", "."))


def _extract_prizes(
    row: dict[str, object], headers: dict[str, str], max_rank: int
) -> tuple[PrizeResult, ...]:
    prizes = []
    for rank in range(1, max_rank + 1):
        winners_column = _find_column(
            headers,
            (f"nombre_de_gagnant_au_rang{rank}", f"nombre_de_gagnant_au_rang_{rank}"),
        )
        payout_column = _find_column(
            headers, (f"rapport_du_rang{rank}", f"rapport_du_rang_{rank}")
        )
        if winners_column is None and payout_column is None:
            continue
        winners = _optional_int(row[winners_column]) if winners_column else None
        payout = _optional_float(row[payout_column]) if payout_column else None
        prizes.append(PrizeResult(rank, winners, payout))
    return tuple(prizes)


def _draws_from_rows(rows: Iterable[dict[str, object]], game: str = "loto") -> list[Draw]:
    rows = list(rows)
    if not rows:
        return []
    original_headers = [str(header) for header in rows[0]]
    headers = {_normalize(header): header for header in original_headers}

    main_columns: list[str] = []
    for position in range(1, 6):
        column = _find_column(
            headers,
            (
                f"boule_{position}",
                f"numero_{position}",
                f"n{position}",
                f"num_{position}",
            ),
        )
        if column is None:
            raise ValueError(f"Colonne du numero principal {position} introuvable")
        main_columns.append(column)

    chance_column = _find_column(
        headers, ("numero_chance", "boule_chance", "chance", "num_chance")
    )
    if chance_column is None:
        raise ValueError("Colonne du numero Chance introuvable")
    date_column = _find_column(headers, ("date_de_tirage", "date_tirage", "date"))

    draws: list[Draw] = []
    for row_number, row in enumerate(rows, start=2):
        try:
            main = tuple(int(float(str(row[column]).replace(",", "."))) for column in main_columns)
            chance = int(float(str(row[chance_column]).replace(",", ".")))
            parsed_date = _parse_date(str(row[date_column])) if date_column else None
            prizes = _extract_prizes(row, headers, 9)
            draws.append(
                Draw(main, chance, parsed_date.date() if parsed_date else None, game, prizes)
            )
        except (TypeError, ValueError, KeyError) as exc:
            raise ValueError(f"Ligne {row_number} invalide: {exc}") from exc
    return draws


def _legacy_draws_from_rows(
    rows: Iterable[dict[str, object]], game: str = "loto"
) -> list[LegacyDraw]:
    rows = list(rows)
    if not rows:
        return []
    headers = {_normalize(str(header)): str(header) for header in rows[0]}
    main_columns = []
    for position in range(1, 7):
        column = _find_column(headers, (f"boule_{position}", f"numero_{position}"))
        if column is None:
            raise ValueError(f"Colonne historique boule_{position} introuvable")
        main_columns.append(column)
    complementary_column = _find_column(
        headers, ("boule_complementaire", "numero_complementaire", "complementaire")
    )
    if complementary_column is None:
        raise ValueError("Colonne du numero complementaire introuvable")
    date_column = _find_column(headers, ("date_de_tirage", "date_tirage", "date"))

    draws = []
    for row_number, row in enumerate(rows, start=2):
        try:
            main = tuple(int(float(str(row[column]))) for column in main_columns)
            complementary = int(float(str(row[complementary_column])))
            parsed_date = _parse_date(str(row[date_column])) if date_column else None
            prizes = _extract_prizes(row, headers, 7)
            draws.append(
                LegacyDraw(
                    main,
                    complementary,
                    parsed_date.date() if parsed_date else None,
                    game,
                    prizes,
                )
            )
        except (TypeError, ValueError, KeyError) as exc:
            raise ValueError(f"Ligne historique {row_number} invalide: {exc}") from exc
    return draws


def _load_csv_bytes(content: bytes, game: str = "loto") -> list[Draw]:
    return _draws_from_rows(_csv_rows(content), game)


def _csv_rows(content: bytes) -> list[dict[str, str]]:
    text = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            pass
    if text is None:
        raise ValueError("Encodage CSV non reconnu")
    header = next((line for line in text.splitlines() if line.strip()), "")
    delimiter = max((";", ",", "\t", "|"), key=header.count)
    if header.count(delimiter) == 0:
        raise ValueError("Separateur CSV non reconnu")
    return list(csv.DictReader(io.StringIO(text), delimiter=delimiter))


def _load_xlsx(path: Path) -> list[Draw]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Installez le support XLSX avec: pip install 'loto-lab[xlsx]'") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(iterator)]
    return _draws_from_rows(dict(zip(headers, row, strict=True)) for row in iterator)


def load_draws(path: str | Path, game: str | None = None) -> list[Draw]:
    source = Path(path)
    if source.suffix.lower() in {".sqlite", ".db"}:
        from .database import load_current_database

        return load_current_database(source)
    source_game = game or _infer_game(source)
    if source.is_dir():
        draws: list[Draw] = []
        for child in sorted(source.iterdir()):
            if child.suffix.lower() in {".csv", ".zip", ".xlsx"}:
                draws.extend(load_draws(child))
        return _deduplicate_and_sort(draws)
    if source.suffix.lower() == ".csv":
        return _deduplicate_and_sort(_load_csv_bytes(source.read_bytes(), source_game))
    if source.suffix.lower() == ".xlsx":
        return _deduplicate_and_sort(_load_xlsx(source))
    if source.suffix.lower() == ".zip":
        draws = []
        with zipfile.ZipFile(source) as archive:
            for name in archive.namelist():
                if name.lower().endswith(".csv"):
                    member_game = (
                        source_game if game or source_game != "loto" else _infer_game(name)
                    )
                    draws.extend(_load_csv_bytes(archive.read(name), member_game))
        if not draws:
            raise ValueError("L'archive ne contient aucun CSV exploitable")
        return _deduplicate_and_sort(draws)
    raise ValueError("Formats acceptes: .csv, .zip, .xlsx, ou un dossier")


def load_draws_many(paths: Iterable[str | Path]) -> list[Draw]:
    return _deduplicate_and_sort(draw for path in paths for draw in load_draws(path))


def load_legacy_draws(
    path: str | Path, from_year: int | None = None, game: str | None = None
) -> list[LegacyDraw]:
    source = Path(path)
    if source.suffix.lower() in {".sqlite", ".db"}:
        from .database import load_legacy_database

        return load_legacy_database(source, from_year)
    source_game = game or _infer_game(source)
    if source.suffix.lower() == ".csv":
        draws = _legacy_draws_from_rows(_csv_rows(source.read_bytes()), source_game)
    elif source.suffix.lower() == ".zip":
        draws = []
        with zipfile.ZipFile(source) as archive:
            for name in archive.namelist():
                if name.lower().endswith(".csv"):
                    member_game = (
                        source_game if game or source_game != "loto" else _infer_game(name)
                    )
                    draws.extend(
                        _legacy_draws_from_rows(_csv_rows(archive.read(name)), member_game)
                    )
    else:
        raise ValueError("L'historique ancien doit etre un CSV ou un ZIP")
    if from_year is not None:
        draws = [
            draw
            for draw in draws
            if draw.draw_date is not None and draw.draw_date.year >= from_year
        ]
    unique = {
        (draw.game, draw.draw_date, draw.main, draw.complementary): draw for draw in draws
    }
    return sorted(
        unique.values(), key=lambda draw: (draw.draw_date is None, draw.draw_date, draw.game)
    )


def load_legacy_draws_many(
    paths: Iterable[str | Path], from_year: int | None = None
) -> list[LegacyDraw]:
    draws = [draw for path in paths for draw in load_legacy_draws(path, from_year)]
    unique = {
        (draw.game, draw.draw_date, draw.main, draw.complementary): draw for draw in draws
    }
    return sorted(
        unique.values(), key=lambda draw: (draw.draw_date is None, draw.draw_date, draw.game)
    )


def _deduplicate_and_sort(draws: Iterable[Draw]) -> list[Draw]:
    unique: dict[tuple[str, object, tuple[int, ...], int], Draw] = {}
    for draw in draws:
        unique[(draw.game, draw.draw_date, draw.main, draw.chance)] = draw
    return sorted(
        unique.values(), key=lambda draw: (draw.draw_date is None, draw.draw_date, draw.game)
    )


def download_latest_archive(destination: str | Path, url: str = LATEST_ARCHIVE_URL) -> Path:
    target = Path(destination)
    target.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(url, headers={"User-Agent": "loto-lab/0.1"})
    with urllib.request.urlopen(request, timeout=30) as response:  # noqa: S310
        content = response.read()
    if not zipfile.is_zipfile(io.BytesIO(content)):
        raise ValueError("La reponse FDJ n'est pas une archive ZIP valide")
    target.write_bytes(content)
    return target


def download_all_archives(destination: str | Path) -> list[Path]:
    directory = Path(destination)
    directory.mkdir(parents=True, exist_ok=True)
    return [
        download_latest_archive(directory / filename, url)
        for filename, url in OFFICIAL_ARCHIVES.items()
    ]
